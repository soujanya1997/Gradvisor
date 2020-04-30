import xlrd
import glob
import os
import xlwt
from xlwt import Workbook
import openpyxl


outputFile1 = "Personal_Consolidation.xls"


wb1 = Workbook()
wb2 = Workbook()
Personal = wb1.add_sheet('Personal_Info')
Technical = wb2.add_sheet('Technical_Info')
outputFile2 = "Technical_Consolidation.xls"
R=0
for root, dirs, files in os.walk(".\\ExcelSheets"):
for file in files:
R = R + 1
f = "ExcelSheets" + "/" +  file
wb = xlrd.open_workbook(f)
sheet2 = wb.sheet_by_name("Sheet2")
sheet3 = wb.sheet_by_name("Sheet3")
# for i in range(11):
#     try:
#         row = sheet2.cell(1, i).value
#         if i == 4:
#             print(row)
#         Personal.write(R,i,row)
#     except:
#         print()
#     wb1.save(outputFile1)




for i in range(7):
try:
row = sheet3.cell_value(2, i)
if(i==0):
Technical.write(R+1,i,row)
else:
str = ""
for j in range(2,100):
try:
row = sheet3.cell_value(j, i)
if(row != ""):
str += row + ", "
except:
print("", end = "")
Technical.write(R+1,i,str)
Personal.write(R,i+10,str)
except:
print(f)


wb1.save(outputFile1)


wb2.save(outputFile2)